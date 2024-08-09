import pandas as pd
from openpyxl import load_workbook

# Path to the Excel file
excel_file_path = 'data/raw/SplitRep_BQ_Unity.xlsx'

# Load the workbook
wb = load_workbook(excel_file_path, read_only=True)

# Get the sheet names (which can represent table names)
sheet_names = wb.sheetnames

# Print the sheet names
print("Table names (sheet names) in the Excel file:")
for sheet in sheet_names:
    print(sheet)

# Function to read data from a specific sheet and save it to a CSV file
def save_sheet_to_csv(sheet_name):
    # Read the sheet into a DataFrame
    df = pd.read_excel(excel_file_path, sheet_name=sheet_name, engine='openpyxl')

    # Check if the DataFrame is empty
    if df.empty:
        print(f"No data found in sheet {sheet_name}.")
    else:
        # Print the DataFrame for verification
        print(f"Data from sheet {sheet_name}:")
        print(df.head())

        # Path to save the CSV file
        csv_file_path = f"data/raw/{sheet_name}.csv"

        # Convert the DataFrame to a CSV file
        df.to_csv(csv_file_path, index=False)

        print(f"Data from {sheet_name} has been successfully saved to {csv_file_path}")

# Example: Read data from a specific sheet and save it as a CSV file
# Replace 'YourSheetName' with the actual sheet name you want to extract
for sheet in sheet_names:
    save_sheet_to_csv(sheet)
